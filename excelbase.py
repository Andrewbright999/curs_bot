import os
import openpyxl

from  config import LESDS_FIELDS

class TableCell:
    def __init__(self, title: str, letter: str) -> None:
        self.title = title
        self.letter = letter


class WorkbookManager:
    """
    Работает непосредственно с excel файлом
    """
    def __init__(self, file_path: str):    
        self.file_path = file_path
        if not os.path.exists(file_path):
            self.workbook = openpyxl.Workbook()
            self.workbook.save(filename=file_path)
        else:
            self.workbook = openpyxl.load_workbook(self.file_path)
        self.delete_sheet()
    
            
    def create_sheet(self, sheet: str = "Sheet"):
        sheets = [sheet.title for sheet in self.workbook]
        if (sheet in sheets) == False:
            self.workbook.create_sheet(sheet)
            self.workbook.save(self.file_path)
            
            
    def delete_sheet(self, sheet: str = "Sheet"):
        sheets = [sheet.title for sheet in self.workbook]
        if sheet in sheets:
            del self.workbook[sheet]
            
    def write_line(self, sheet: str = "Sheet", row_number: int = None, search_by_letter: str = "A",  **kwargs):
        sheet = self.workbook[sheet]
        if row_number:
            pass
        else:    
            for cell in sheet[search_by_letter]:
                print(f"Координаты: {cell.coordinate} Значение: {cell.value}")
                if cell.value is None:
                    row_number = cell.row
                    break
        
        print(row_number)
            # first_unfilled_cell.value = kwargs.key
    
    
        

class SheetManager:
    """
    Отвечает за работу с листами таблицы, полями на ней и тп
    """
    def __init__(self, table: WorkbookManager, sheet_name, fields_config):
        self.table = table
        self.sheet = sheet_name
        self.table.create_sheet(sheet = self.sheet)
        self.table.write_line(sheet = self.sheet)
        self.sheet_name = sheet_name
        for field in fields_config:
            cell = fields_config[f"{field}"]
            setattr(self, field, TableCell(title = cell["title"], letter = cell["letter"]))
            self.table.write_line(sheet = self.sheet)

table = WorkbookManager(file_path = "sheeet.xlsx")
lead_sheet = SheetManager(table = table, sheet_name = "лиды нахуй", fields_config = LESDS_FIELDS )



# sheet = SheetList()


# workbook = Workbook()
# print(workbook.id.column_title)


        

leadbook_path = "leads.xlsx"

# leadbook = Workbook_Manager(leadbook_path)

# print(leadbook)

# leadbook.write_user(telegram="G")



leads_json_example = [{
    "id": 121232132,
    "name": "Андрей",
    "email": None,
    "telegram": "Bright099",
    "check_in": True,
    "link": "https://t.me/Bright099",
    "block_bot": False
    },
    {
    "id": 121232132,
    "name": "Андрей",
    "email": None,
    "telegram": "Bright099",
    "check_in": True,
    "link": "https://t.me/Bright099",
    "block_bot": False
    },
]

# for user in fields_config_json:
# for row in leads_json_example:
#     print(fields_config_json[row]["letter"])

# for i in fields_config_json:
#     print(fields_config_json[i]["title"])

# cell_val = sheet["A1"].value
# print(cell_val)
# cell_val = sheet.cell(row=2, column=1).value
# print(cell_val)


